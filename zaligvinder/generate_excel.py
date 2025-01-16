import openpyxl
import os
import argparse
import storage.sqlitedb
import re

def getInstanceIds(db):
    trackinstance_db = storage.sqlitedb.TrackInstanceRepository(db)
    track_db = storage.sqlitedb.TrackRepository(db, trackinstance_db)
    result_db = storage.sqlitedb.ResultRepository(
        db, track_db, trackinstance_db)
    trackids = result_db.getTrackIds()
    # sort the instances by file path
    instanceIds = reversed(
        sorted(
            [(instanceid, result_db.getFilePath(instanceid))
                for trackid in trackids 
                for instanceid in result_db.getInstanceIdsForTrack(trackid)
            ],
            key=lambda x: x[1]
        )
    )
    instanceIds = [instanceid for instanceid, _ in instanceIds]
    return instanceIds

def getResult(db, instanceid, solvers):
    trackinstance_db = storage.sqlitedb.TrackInstanceRepository(db)
    track_db = storage.sqlitedb.TrackRepository(db, trackinstance_db)
    result_db = storage.sqlitedb.ResultRepository(
        db, track_db, trackinstance_db)
    result = result_db.getInstanceResultForSolvers(instanceid,solvers)
    # the final result format [path, solver1 (res, time), solver2 (res, time), ...]
    path = result_db.getFilePath(result[0][1])
    path = path[path.find("cav2025")+len("cav2025"):]
    # sort the result by solver name
    result = sorted(result, key=lambda x: x[0])
    finalres = [(solverRes[4].result, solverRes[4].time/1000) for solverRes in result]
    # result  1 is sat, 0 is unsat, None is unknown or timeout
    finalres = [path] + [("sat", res[1]) if res[0] == 1 else ("unsat", res[1]) if res[0] == 0 else ("", res[1]) for res in finalres]
    print(finalres)
    return finalres


def getExcelData(db):
    trackinstance_db = storage.sqlitedb.TrackInstanceRepository(db)
    track_db = storage.sqlitedb.TrackRepository(db, trackinstance_db)
    result_db = storage.sqlitedb.ResultRepository(
        db, track_db, trackinstance_db)
    # sort the solvers by name
    solvers = sorted(result_db.getSolvers())
    
    data = [["File Name"] + [f"{solver}" for solver in solvers]]
    instanceIds = getInstanceIds(db)
    for instanceid in instanceIds:
        row = getResult(db, instanceid, solvers)
        data.append(row)
    
    emptyrow = [""]
    # sat files for each solver
    satfiles = [""] + [("#sat", result_db.getSummaryForSolver(solver)[2]) for solver in solvers]
    # unsat files for each solver
    unsatfiles = [""] + [("#unsat", result_db.getSummaryForSolver(solver)[4]) for solver in solvers]
    # time for each solver
    time = [""] + [("#Time(s)", round(result_db.getSummaryForSolver(solver)[6]/1000/result_db.getSummaryForSolver(solver)[7], 2)) for solver in solvers]
    data.append(emptyrow)
    data.append(satfiles)
    data.append(unsatfiles)
    data.append(time)
    
    data.append(emptyrow)
    data.append(["", ("#Total", len(instanceIds))])
    
    # print summary
    for solver in solvers:
        summay = result_db.getSummaryForSolver(solver)
        print(f"{solver}: sat: {summay[2]}, unsat: {summay[4]}, time: {summay[6]/1000/summay[7]}")
    return data


argparser = argparse.ArgumentParser(
    prog=__file__, description="print the error instances of a database")
argparser.add_argument("database")
args = argparser.parse_args()
db = storage.sqlitedb.DB(args.database)



# Data example
data = [
    ["File Name", "Solver 1 (res, time)", "Solver 2 (res, time)", "Solver 3 (res, time)", 
     "Solver 4 (res, time)", "Solver 5 (res, time)", "Solver 6 (res, time)"],
    ["file1", (0.1, 1.2), (0.2, 1.4), (0.3, 1.3), (0.4, 1.1), (0.5, 1.5), (0.6, 1.6)],
    ["file2", (0.2, 2.1), (0.3, 2.2), (0.4, 2.3), (0.5, 2.4), (0.6, 2.5), (0.7, 2.6)],
    # Add more rows as needed
]

data = getExcelData(db)

# Prepare the data for Excel
formatted_data = []

# Add headers
header_row = ["File Name"]
for solver in data[0][1:]:
    header_row.append(solver)  # Keep the title as is
    header_row.append("")  # Placeholder for the second column of the solver
formatted_data.append(header_row)

# Add subheaders for Res and Time
subheader_row = [""]  # First column is empty
for solver in data[0][1:]:
    subheader_row.extend(["Res", "Time"])
formatted_data.append(subheader_row)

# Add data rows
for row in data[1:]:
    formatted_row = [row[0]]  # Start with the file name
    for cell in row[1:]:
        if isinstance(cell, tuple):  # Separate res and time
            formatted_row.extend(cell)
        else:
            formatted_row.extend([cell, ""])
    formatted_data.append(formatted_row)

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Results"

# Write formatted data to the worksheet row by row
for row_idx, row in enumerate(formatted_data, start=1):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Merge cells for solver titles (first row)
for col_index in range(2, len(header_row), 2):  # Start at column 2 and merge every 2 columns
    ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + 1)

# Save the workbook to an Excel file
output_file = "solver_results.xlsx"
wb.save(output_file)

print(f"Excel file '{output_file}' created successfully.")
